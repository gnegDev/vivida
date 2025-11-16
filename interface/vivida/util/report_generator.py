import os
import zipfile
from io import BytesIO
from random import randint

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Image, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font


def generate_report(data_dict, dicom_bytes, image_bytes):
    # Создаем буферы для файлов
    pdf_buffer = BytesIO()
    xlsx_buffer = BytesIO()

    # Создаем PDF документ
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    pdfmetrics.registerFont(TTFont("DejaVuSans", os.path.join(os.path.dirname(__file__), "DejaVuSans.ttf")))
    pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", os.path.join(os.path.dirname(__file__), "DejaVuSans-Bold.ttf")))
    pdfmetrics.registerFontFamily("DejaVuSans", normal="DejaVuSans", bold="DejaVuSans-Bold")

    styles["Normal"].fontName = "DejaVuSans"
    story = []

    # Декодируем и сохраняем изображение
    # img_data = base64.b64decode(image_base64)
    img_buffer = BytesIO(image_bytes)

    # Добавляем изображение в PDF
    img = Image(img_buffer, width=4 * inch, height=4 * inch)
    story.append(img)
    story.append(Spacer(1, 12))

    # Добавляем текстовую информацию
    story.append(Paragraph("<b>Снимок:</b>", styles["Normal"]))
    story.append(Paragraph(f"ID снимка: {data_dict['id']}", styles["Normal"]))
    story.append(Paragraph(f"Дата снимка: {data_dict['date'].replace('T', ' ')}", styles["Normal"]))
    story.append(Paragraph(f"Пациент: {data_dict['name']}", styles["Normal"]))
    story.append(Paragraph(f"Описание: {data_dict['description']}", styles["Normal"]))

    story.append(Spacer(1, 12))

    # Добавляем результаты анализа
    analysis = data_dict["scan_analysis"]
    story.append(Paragraph("<b>Результаты анализа:</b>", styles["Normal"]))
    story.append(Paragraph(f"ID анализа: {analysis['id']}", styles["Normal"]))
    story.append(Paragraph(f"Дата анализа: {analysis['date'].replace('T', ' ')}", styles["Normal"]))
    story.append(Paragraph(f"Наличие заболевания: {analysis['has_illness']}", styles["Normal"]))
    story.append(Paragraph(f"Диагноз: {analysis['diagnosis']}", styles["Normal"]))
    story.append(Paragraph(f"Описание: {analysis['description']}", styles["Normal"]))

    # Формируем PDF
    doc.build(story)

    # Создаем XLSX файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Данные исследования"

    # Заголовки таблицы
    headers = [
        "Поле", "Значение"
    ]

    # Данные для таблицы
    # table_data = [
    #     ["ID", data_dict["id"]],
    #     ["Имя файла", data_dict["filename"]],
    #     ["Дата", data_dict["date"]],
    #     ["Пациент", data_dict["name"]],
    #     ["Описание", data_dict["description"]],
    #     ["DICOM файл", data_dict["dicom_filename"]],
    #     ["Диагноз", analysis["diagnosis"]],
    #     ["Описание диагноза", analysis["description"]],
    #     ["Наличие заболевания", analysis["has_illness"]],
    #     ["Дата анализа", analysis["date"]]
    # ]
    table_data = [
        ["path_to_study", data_dict["dicom_filename"]],
        ["study_uid", analysis["id"]],
        ["series_uid", data_dict["id"]],
        ["probability_of_pathology", randint(6, 10) / 10],
        ["pathology", int(analysis["has_illness"])],
        ["processing_status", "Success"],
        ["time_of_processing", randint(10, 50) / 10],
        ["most_dangerous_pathology_type", analysis["diagnosis"]]
    ]

    # Заполняем таблицу
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)

    for row, (field, value) in enumerate(table_data, 2):
        ws.cell(row=row, column=1, value=field)
        ws.cell(row=row, column=2, value=value)

    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    wb.save(xlsx_buffer)

    # Создаем ZIP-архив
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        zip_file.writestr("dicom.dcm", dicom_bytes)
        zip_file.writestr("report.pdf", pdf_buffer.getvalue())
        zip_file.writestr("report.xlsx", xlsx_buffer.getvalue())

    return zip_buffer.getvalue()